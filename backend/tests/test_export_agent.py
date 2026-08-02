import os
import pytest  # type: ignore # pylint: disable=import-error
import openpyxl
from unittest.mock import AsyncMock, MagicMock, patch
from bson.objectid import ObjectId  # type: ignore # pylint: disable=import-error

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

from agents.export_agent import (
    get_export_filename,
    get_jobs_by_ids,
    generate_excel_report,
    generate_pdf_report
)

def test_get_export_filename():
    job_ids = [str(ObjectId()), str(ObjectId())]
    # Test deterministic sorting
    filename1 = get_export_filename(job_ids, "excel")
    filename2 = get_export_filename(list(reversed(job_ids)), "excel")
    
    assert filename1 == filename2
    assert filename1.endswith(".xlsx")
    
    filename_pdf = get_export_filename(job_ids, "pdf")
    assert filename_pdf.endswith(".pdf")
    assert filename_pdf != filename1

@pytest.mark.asyncio
@patch("agents.export_agent.get_database")
async def test_get_jobs_by_ids_fallback(mock_get_db):
    mock_db = MagicMock()
    mock_get_db.return_value = mock_db
    
    job_id = ObjectId()
    # Job lacking optional fields
    mock_db.jobs.find_one = AsyncMock(return_value={
        "_id": job_id,
        "company": "Test Company",
        "role": "Software Engineer"
    })
    
    jobs = await get_jobs_by_ids([str(job_id)])
    assert len(jobs) == 1
    assert jobs[0]["company"] == "Test Company"
    assert jobs[0]["location"] == "N/A"
    assert jobs[0]["salary"] == "N/A"
    assert jobs[0]["apply_link"] == "N/A"
    assert jobs[0]["matched_skills"] == []
    assert jobs[0]["missing_skills"] == []
    assert jobs[0]["skill_match_score"] is None

def test_generate_excel_report(tmp_path):
    filepath = str(tmp_path / "test_report.xlsx")
    jobs = [
        {
            "company": "Company A",
            "role": "Role A",
            "location": "Loc A",
            "salary": "Sal A",
            "relevance_score": 95,
            "skill_match_score": 80,
            "matched_skills": ["Python", "SQL"],
            "missing_skills": ["Docker"],
            "apply_link": "http://linka.com"
        },
        {
            "company": "Company B",
            "role": "Role B",
            "location": "Loc B",
            "salary": "Sal B",
            "relevance_score": None,
            "skill_match_score": None,
            "matched_skills": [],
            "missing_skills": [],
            "apply_link": "N/A"
        }
    ]
    
    generate_excel_report(jobs, filepath)
    assert os.path.exists(filepath)
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    assert ws.title == "Jobs Export"
    
    # Headers check
    headers = [cell.value for cell in ws[1]]
    assert "Company" in headers
    assert "Skill Match Score" in headers
    
    # Check populated data
    row2 = [cell.value for cell in ws[2]]
    assert row2[0] == "Company A"
    assert row2[4] == 95
    assert row2[5] == 80
    assert row2[6] == "Python, SQL"
    assert row2[7] == "Docker"
    assert row2[8] == "http://linka.com"

    row3 = [cell.value for cell in ws[3]]
    assert row3[0] == "Company B"
    assert row3[4] == "N/A"
    assert row3[5] == "N/A"
    assert row3[6] == "None"
    assert row3[7] == "None"
    assert row3[8] == "N/A"

def test_generate_pdf_report(tmp_path):
    filepath = str(tmp_path / "test_report.pdf")
    jobs = [
        {
            "company": "Company A",
            "role": "Role A",
            "location": "Loc A",
            "salary": "Sal A",
            "relevance_score": 95,
            "skill_match_score": 80,
            "matched_skills": ["Python", "SQL"],
            "missing_skills": ["Docker"],
            "apply_link": "http://linka.com"
        }
    ]
    
    generate_pdf_report(jobs, filepath)
    assert os.path.exists(filepath)
    # Basic verification of PDF header bytes
    with open(filepath, "rb") as f:
        header = f.read(5)
    assert header == b"%PDF-"
