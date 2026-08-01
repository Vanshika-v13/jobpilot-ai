import os
import logging
import hashlib
from datetime import datetime
from typing import List, Dict, Any
from bson import ObjectId

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from database.connection import get_database

logger = logging.getLogger(__name__)

def get_export_filename(job_ids: List[str], format_ext: str) -> str:
    """
    Generate a deterministic filename by hashing the sorted list of job IDs and the format.
    """
    sorted_ids = sorted([str(jid) for jid in job_ids])
    data_str = f"{','.join(sorted_ids)}:{format_ext}"
    file_hash = hashlib.md5(data_str.encode("utf-8")).hexdigest()
    ext = "xlsx" if format_ext == "excel" else "pdf"
    return f"{file_hash}.{ext}"

async def get_jobs_by_ids(job_ids: List[str]) -> List[Dict[str, Any]]:
    """
    Fetch job documents from MongoDB. Fallback missing optional fields to default placeholders.
    """
    db = get_database()
    jobs = []
    for jid in job_ids:
        try:
            oid = ObjectId(jid)
            job = await db.jobs.find_one({"_id": oid})
            if job:
                # Standardize job dict fields
                job["_id"] = str(job["_id"])
                # Fallback fields
                job["company"] = job.get("company") or "N/A"
                job["role"] = job.get("role") or "N/A"
                job["location"] = job.get("location") or "N/A"
                job["salary"] = job.get("salary") or "N/A"
                job["apply_link"] = job.get("apply_link") or "N/A"
                job["skill_match_score"] = job.get("skill_match_score")
                job["matched_skills"] = job.get("matched_skills") or []
                job["missing_skills"] = job.get("missing_skills") or []
                job["relevance_score"] = job.get("relevance_score")
                jobs.append(job)
            else:
                logger.warning(f"Job with ID {jid} not found in database.")
        except Exception as e:
            logger.error(f"Error fetching job {jid}: {e}")
    return jobs

def generate_excel_report(jobs: List[Dict[str, Any]], filepath: str) -> None:
    """
    Generate an Excel sheet of job listings using openpyxl.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs Export"

    headers = [
        "Company", "Role", "Location", "Salary", "Relevance Score", 
        "Skill Match Score", "Matched Skills", "Missing Skills", "Apply Link"
    ]
    ws.append(headers)

    # Style Header Row
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Populate Data Rows
    for row_idx, job in enumerate(jobs, 2):
        matched = ", ".join(job.get("matched_skills") or [])
        missing = ", ".join(job.get("missing_skills") or [])
        
        relevance_score = job.get("relevance_score")
        if relevance_score is None:
            relevance_score = "N/A"
            
        skill_match_score = job.get("skill_match_score")
        if skill_match_score is None:
            skill_match_score = "N/A"

        row_data = [
            job.get("company"),
            job.get("role"),
            job.get("location"),
            job.get("salary"),
            relevance_score,
            skill_match_score,
            matched or "None",
            missing or "None",
            job.get("apply_link")
        ]
        
        ws.append(row_data)
        
        # Style Data Cells
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            
            # Apply link cell formatting
            if col_idx == len(headers) and job.get("apply_link") and job.get("apply_link") != "N/A":
                cell.hyperlink = job.get("apply_link")
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            
            # Center alignment for scores
            if col_idx in [5, 6]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.hyperlink:
                max_len = max(max_len, min(len(val_str), 30))
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    wb.save(filepath)

def generate_pdf_report(jobs: List[Dict[str, Any]], filepath: str) -> None:
    """
    Generate a formatted PDF document using reportlab.platypus.
    """
    doc = SimpleDocTemplate(
        filepath,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        leading=28,
        textColor=colors.HexColor('#1F4E79'),
        spaceAfter=15
    )
    
    job_title_style = ParagraphStyle(
        'JobTitle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#2E75B6'),
        spaceBefore=10,
        spaceAfter=5
    )
    
    normal_style = ParagraphStyle(
        'NormalStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#333333')
    )
    
    bold_label_style = ParagraphStyle(
        'BoldLabel',
        parent=normal_style,
        fontName='Helvetica-Bold'
    )

    link_style = ParagraphStyle(
        'LinkStyle',
        parent=normal_style,
        textColor=colors.HexColor('#0563C1'),
        fontName='Helvetica-Oblique'
    )

    story = []
    
    # Document Header
    story.append(Paragraph("JobPilot AI - Job Search Report", title_style))
    story.append(Paragraph(f"Compiled on {datetime.now().strftime('%Y-%m-%d %H:%M')}", normal_style))
    story.append(Spacer(1, 15))
    
    for i, job in enumerate(jobs):
        job_story = []
        
        # Job title / header
        role = job.get("role") or "Unknown Role"
        company = job.get("company") or "Unknown Company"
        job_story.append(Paragraph(f"{i+1}. {role} @ {company}", job_title_style))
        
        # Metadata table
        loc = job.get("location") or "N/A"
        sal = job.get("salary") or "N/A"
        link = job.get("apply_link") or "N/A"
        match_score = job.get("skill_match_score")
        match_score_str = f"{match_score}%" if match_score is not None else "N/A"
        rel_score = job.get("relevance_score")
        rel_score_str = str(rel_score) if rel_score is not None else "N/A"

        if link and link != "N/A":
            link_p = Paragraph(f'<a href="{link}">Apply Link</a>', link_style)
        else:
            link_p = Paragraph("N/A", normal_style)

        table_data = [
            [Paragraph("Location:", bold_label_style), Paragraph(loc, normal_style),
             Paragraph("Salary:", bold_label_style), Paragraph(sal, normal_style)],
            [Paragraph("Skill Match:", bold_label_style), Paragraph(match_score_str, normal_style),
             Paragraph("Relevance Score:", bold_label_style), Paragraph(rel_score_str, normal_style)],
            [Paragraph("Link:", bold_label_style), link_p,
             Paragraph("", bold_label_style), Paragraph("", normal_style)]
        ]
        
        t = Table(table_data, colWidths=[80, 180, 100, 170])
        t.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        job_story.append(t)
        job_story.append(Spacer(1, 6))

        # Skills details
        matched_list = job.get("matched_skills") or []
        missing_list = job.get("missing_skills") or []
        
        matched_str = ", ".join(matched_list) if matched_list else "None"
        missing_str = ", ".join(missing_list) if missing_list else "None"
        
        job_story.append(Paragraph(f"<b>Matched Skills:</b> {matched_str}", normal_style))
        job_story.append(Spacer(1, 4))
        job_story.append(Paragraph(f"<b>Missing Skills:</b> {missing_str}", normal_style))
        
        # Divider Line
        job_story.append(Spacer(1, 10))
        divider = Table([[""]], colWidths=[530])
        divider.setStyle(TableStyle([
            ('LINEBELOW', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ]))
        job_story.append(divider)
        job_story.append(Spacer(1, 10))

        story.append(KeepTogether(job_story))
        
    doc.build(story)
